import pandas as pd
import os
import warnings
warnings.filterwarnings('ignore')
import glob
import openpyxl


path =r'\\prod-app02\ConvenienceCheckApproval\WellsFargoFiles' # use your path
allFiles = glob.glob(path + "/*.xlsx")
bank = pd.DataFrame()
list_ = []
for file_ in allFiles:
    bank = pd.read_excel(file_,index_col=None, header=0,
                         encoding="ISO-8859-1", error_bad_lines=False)
    list_.append(bank)
bank = pd.concat(list_)
bank.columns = bank.columns.str.strip().str.lower().str.replace(' ', '_')
bank.columns = bank.columns.str.strip().str.lower().str.replace('-', '_')

bank['customer_ref_no'] = bank['customer_ref_no'].astype(str) 
bank['customer_ref_no'] = bank['customer_ref_no'].apply(lambda x: '{0:0>10}'.format(x))

updates = dict(zip(bank.customer_ref_no, bank.as_of_date))

for i in os.listdir(os.chdir(r'M:\2019 Programs\Master Spreadsheets\NB_MASTER_DO_NOT_OPEN')):
    if i.endswith(".xlsx"):
        wb = openpyxl.load_workbook(i)
        sheet = wb.get_sheet_by_name('Setupheader')
        for rowNum in range(2, sheet.max_row):  # skip the first row
            checkNumber = sheet.cell(row=rowNum, column=19).value
            if checkNumber in updates:
                sheet.cell(row=rowNum, column=20).value = updates[checkNumber]
        wb.save(i)

for i in os.listdir(os.chdir(r'M:\2019 Programs\Master Spreadsheets\FB_MASTER_DO_NOT_OPEN')):
    if i.endswith(".xlsx"):
        wb = openpyxl.load_workbook(i)
        sheet = wb.get_sheet_by_name('Setupheader')
        for rowNum in range(2, sheet.max_row):  # skip the first row
            checkNumber = sheet.cell(row=rowNum, column=19).value
            if checkNumber in updates:
                sheet.cell(row=rowNum, column=20).value = updates[checkNumber]
        wb.save(i)


import pyodbc

# Parameters
server = 'NLS-Prod-SQL03'
db = 'NLS_Prod'

# Create the connection
conn = pyodbc.connect('DRIVER={SQL Server};SERVER=' + server + ';DATABASE=' + db + ';Trusted_Connection=yes')

# query db
sql = '''
SELECT
    t1.acctrefno, 
    t1.cifno, 
    t1.name, 
    t1.loan_number, 
    t1.input_date,
    t1.open_date, 
    t2.userdef10 AS checknumber

FROM [NLS_Prod].[dbo].[loanacct] t1

INNER JOIN [NLS_Prod].[dbo].[loanacct_detail] t2 
ON t1.acctrefno = t2.acctrefno
WHERE t2.userdef10 IS NOT NULL
'''
loanacct = pd.io.sql.read_sql(sql, conn)

loanacct['acctrefno'] = loanacct['acctrefno'].fillna(0).astype(int)
loanacct['cifno'] = loanacct['cifno'].fillna(0).astype(int)

updates = dict(zip(loanacct.checknumber, loanacct.loan_number))

for i in os.listdir(os.chdir(r'M:\2019 Programs\Master Spreadsheets\NB_MASTER_DO_NOT_OPEN')):
    if i.endswith(".xlsx"):
        wb = openpyxl.load_workbook(i)
        sheet = wb.get_sheet_by_name('Setupheader')
        for rowNum in range(2, sheet.max_row):  # skip the first row
            checkNumber = sheet.cell(row=rowNum, column=19).value
            if checkNumber in updates:
                sheet.cell(row=rowNum, column=21).value = updates[checkNumber]
        wb.save(i)
        
for i in os.listdir(os.chdir(r'M:\2019 Programs\Master Spreadsheets\FB_MASTER_DO_NOT_OPEN')):
    if i.startswith("FB_Cleared"):
        wb = openpyxl.load_workbook(i)
        sheet = wb.get_sheet_by_name('Setupheader')
        for rowNum in range(2, sheet.max_row):  # skip the first row
            checkNumber = sheet.cell(row=rowNum, column=19).value
            if checkNumber in updates:
                sheet.cell(row=rowNum, column=21).value = updates[checkNumber]
        wb.save(i)

import shutil
import os

source = 'M:\\\\2019 Programs\\\\Master Spreadsheets\\\\NB_MASTER_DO_NOT_OPEN\\\\'
dest1 = 'M:\\\\2019 Programs\\\\Master Spreadsheets\\\\'
files = os.listdir(source)
for f in files:
        shutil.copy(source+f, dest1)

source = 'M:\\\\2019 Programs\\\\Master Spreadsheets\\\\FB_MASTER_DO_NOT_OPEN\\\\'
dest1 = 'M:\\\\2019 Programs\\\\'
files = os.listdir(source)
for f in files:
        shutil.copy(source+f, dest1)