import pandas as pd
import os
import warnings
warnings.filterwarnings('ignore')
import glob
import openpyxl


path =r'\\prod-app02\ConvenienceCheckApproval\WellsFargoFiles\2018' # use your path
allFiles = glob.glob(path + "/*.xlsx")
bank = pd.DataFrame()
list_ = []
for file_ in allFiles:
    bank = pd.read_excel(file_,index_col=None, header=0,
                         encoding="ISO-8859-1", error_bad_lines=False)
    list_.append(bank)
bank = pd.concat(list_)
bank.columns = bank.columns.str.strip().str.lower().str.replace(' ', '_')
bank.columns = bank.columns.str.strip().str.lower().str.replace('-', '_')

bank['customer_ref_no'] = bank['customer_ref_no'].astype(str) 
bank['customer_ref_no'] = bank['customer_ref_no'].apply(lambda x: '{0:0>10}'.format(x))

updates = dict(zip(bank.customer_ref_no, bank.as_of_date))

for i in os.listdir(os.chdir(r'E:\cepps\NB\NB_2018_CCMasters')):
    if i.endswith(".xlsx"):
        wb = openpyxl.load_workbook(i)
        sheet = wb.get_sheet_by_name('Setupheader')
        for rowNum in range(2, sheet.max_row):  # skip the first row
            checkNumber = sheet.cell(row=rowNum, column=19).value
            if checkNumber in updates:
                sheet.cell(row=rowNum, column=20).value = updates[checkNumber]
        wb.save(i)